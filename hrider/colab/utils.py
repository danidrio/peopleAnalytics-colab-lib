import os
import pandas as pd
import warnings
from openpyxl import load_workbook

DRIVE_USER_PATH = "/content/drive"
MY_DRIVE_PATH = DRIVE_USER_PATH + "/My Drive/"
_IS_COLAB = None

def is_colab():
    global _IS_COLAB

    if _IS_COLAB is None:
        try:
            import google.colab
            _IS_COLAB = True
        except ImportError:
            _IS_COLAB = False

    return _IS_COLAB

def mount_google_drive():
    if not is_colab():
            raise RuntimeError("Google Drive solo se puede montar automáticamente en Google Colab")

    from google.colab import drive
    if not os.path.exists(DRIVE_USER_PATH):
        drive.mount(DRIVE_USER_PATH)
  
def get_drive_path(filename):
    return os.path.join(MY_DRIVE_PATH, filename)

def add_suffix_to_path(input_path, suffix):
    base, ext = os.path.splitext(input_path)
    return f"{base}_{suffix}{ext}"


def get_secret(name): 
    """
    Retorna un secreto del usuario en Colab. En local el secreto se busca como 
    variable de entorno.
    """
    if is_colab():
        from google.colab import userdata
        key = userdata.get(name)
        if not key:
            raise ValueError(
                f"No has configurado {name}. "
                "Abre el panel Secrets de Colab, crea un secreto llamado {name} "
                "y activa el acceso para este notebook."
            )
        else:
            return key
    else:
        key = os.environ.get(name)
        if not key:
            raise ValueError(f"Secreto {name} no especificado como variable de entorno")
        else:
            return key
    
def read_excel_from_drive(filename):
    # Suppress the specific openpyxl UserWarning
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        file_path = get_drive_path(filename)
        return pd.read_excel(file_path)

def read_excel(filename):
    return pd.read_excel(filename)


def load_people_from_excel (
    excel_path,
    sheet_name=None,
    skip_first_row=True
):
    """
    Carga personas desde un Excel sin depender de los nombres de columna.

    Estructura esperada:
    - Columna 1: employee_id
    - Columna 2: name
    - Columna 3: lastname
    - Columna 4: email, opcional

    Si skip_first_row=True, ignora la primera fila.
    """
    workbook = load_workbook(
        filename=str(excel_path),
        read_only=True,
        data_only=True
    )

    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    people = []

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if skip_first_row and row_index == 1:
            continue

        employee_id = row[0] if len(row) > 0 else None
        name = row[1] if len(row) > 1 else None
        lastname = row[2] if len(row) > 2 else None
        email = row[3] if len(row) > 3 else None

        employee_id = str(employee_id or "").strip()
        name = str(name or "").strip()
        lastname = str(lastname or "").strip()
        email = str(email or "").strip()

        if not employee_id and not name and not lastname and not email:
            continue

        people.append({
            "employee_id": employee_id,
            "name": name,
            "lastname": lastname,
            "email": email
        })

    workbook.close()
    return people